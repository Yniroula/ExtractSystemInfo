import socket
import uuid
import platform
import datetime
import openpyxl

# 
def get_system_info():
    system = platform.system()
    version = platform.version()
    mac_address = ':'.join(['{:02x}'.format((uuid.getnode() >> ele) & 0xff) for ele in range(0,8*6,8)][::-1])
    ip_address = socket.gethostbyname(socket.gethostname())
    machine_name = socket.gethostname()
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return {
        'system': system,
        'version': version,
        'mac_address': mac_address,
        'ip_address': ip_address,
        'machine_name': machine_name,
        'timestamp': timestamp
    }

system_info = get_system_info()

# Creating a new Excel file if it does not exist, or opening it in append mode if it exist

try:
    wb = openpyxl.load_workbook('system_file.xlsx')
except FileNotFoundError:
    wb = openpyxl.Workbook()

ws = wb.create_sheet()

# Add the system information to the Excel file
ws['A1'] = 'System Name'
ws['B1'] = 'System Version'
ws['C1'] = 'MAC Address'
ws['D1'] = 'IP Address'
ws['E1'] = 'Machine Name'
ws['F1'] = 'Timestamp'
ws['A2'] = system_info['system_name']
ws['B2'] = system_info['system_version']
ws['C2'] = system_info['mac_address']
ws['D2'] = system_info['ip_address']
ws['E2'] = system_info['machine_name']
ws['F2'] = system_info['timestamp']


# Saving the changes inthe Excel file
wb.save('system_file.xlsx')